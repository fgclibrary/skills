from __future__ import annotations

import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


ACTION_MAP = [
    ("Reset", "重置"),
    ("Delete", "删除"),
    ("Remove", "移除"),
    ("Update", "更新"),
    ("Edit", "编辑"),
    ("Add", "新增"),
    ("Assign", "分配"),
    ("Combine", "合并"),
    ("reorganize", "调整"),
]

ENTITY_MAP = {
    "Org": "组织",
    "Orgs": "组织",
    "Account": "账户",
    "Accounts": "账户",
    "Role": "角色",
    "RoleType": "角色分类",
    "Members": "成员",
    "Member": "成员",
    "Password": "密码",
    "Users": "用户",
    "User": "用户",
    "PermissionGroup": "权限组",
    "Group": "组",
}

NAME_OVERRIDES = {
    "AddOrg": "新增组织",
    "UpdateOrg": "更新组织",
    "reorganizeOrg": "调整组织层级",
    "DeleteOrgs": "删除组织",
    "AddOrgMembers": "添加组织成员",
    "RemoveMemberFormOrg": "移除组织成员",
    "AddAccount": "新增账户",
    "DeleteAccount": "删除账户",
    "EditAccount": "编辑账户",
    "ResetPassword": "重置账户密码",
    "assignRolesToUsers": "分配角色到用户",
    "assignOrgToUsers": "分配部门到用户",
    "AddRole": "新增角色",
    "EditRole": "编辑角色",
    "DeleteRole": "删除角色或角色分类",
    "AddRoleType": "新增角色分类",
    "EditRoleType": "编辑角色分类",
    "reorganizeRole": "调整角色结构",
    "AddRoleMember": "添加角色成员",
    "RemoveMembersFromRole": "移除角色成员",
    "AssignRoleToOrg": "绑定角色到组织",
    "CombineRoleGroup": "分配权限角色到角色",
    "CombineRolePermissionGroup": "调整角色权限组",
}


@dataclass
class ArrayDetail:
    name: str
    items: list[str] = field(default_factory=list)


@dataclass
class ApiDoc:
    name: str
    title: str
    display_name: str
    method: str
    path: str
    summary: str
    params: list[dict[str, str]]
    array_details: dict[str, ArrayDetail] = field(default_factory=dict)


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def safe_cell(ws, row: int, col: int) -> str:
    return normalize_text(ws.cell(row, col).value)


def guess_display_name(name: str, title: str, path: str) -> str:
    if name in NAME_OVERRIDES:
        return NAME_OVERRIDES[name]

    action = ""
    for prefix, translated in ACTION_MAP:
        if name.startswith(prefix):
            action = translated
            break

    segments = re.findall(r"[A-Z]?[a-z]+|[A-Z]+(?=[A-Z]|$)", name)
    entity_parts = [ENTITY_MAP.get(segment, "") for segment in segments[1:]]
    entity = "".join(part for part in entity_parts if part)

    if not entity:
        path_tail = path.rstrip("/").split("/")[-1] if path else ""
        entity = ENTITY_MAP.get(path_tail, "") or title or name

    guessed = f"{action}{entity}".strip()
    return guessed or title or name


def build_schema_for_param(api: ApiDoc, param: dict[str, str]) -> dict:
    raw_type = param["type"]
    name = param["name"]
    description = param["desc"] or ""

    if raw_type == "数组类型":
        detail = api.array_details.get(name)
        if detail and detail.items:
            properties = {
                item_name: {
                    "type": "string",
                    "description": f"{name} 数组项字段",
                }
                for item_name in detail.items
            }
            return {
                "type": "array",
                "description": description or f"{name} 数组参数",
                "items": {
                    "type": "object",
                    "properties": properties,
                    "required": list(properties.keys()),
                    "additionalProperties": False,
                },
            }
        return {
            "type": "array",
            "description": description or f"{name} 数组参数",
            "items": {"type": "string"},
        }

    return {
        "type": "string",
        "description": description or f"{name} 参数",
    }


def parse_interfaces(source_xlsx: Path) -> list[ApiDoc]:
    wb = load_workbook(source_xlsx, data_only=True)
    list_ws = wb[wb.sheetnames[0]]
    detail_ws = wb[wb.sheetnames[1]]

    interface_list: list[tuple[str, str]] = []
    for row in range(2, list_ws.max_row + 1):
        name = safe_cell(list_ws, row, 4)
        title = safe_cell(list_ws, row, 6)
        if name:
            interface_list.append((name, title))

    starts: dict[str, int] = {}
    for row in range(1, detail_ws.max_row + 1):
        candidate = safe_cell(detail_ws, row, 1)
        if candidate and re.match(r"^[A-Za-z][A-Za-z0-9_]+$", candidate):
            starts.setdefault(candidate, row)

    sorted_starts = sorted(starts.items(), key=lambda item: item[1])
    next_row_by_name: dict[str, int] = {}
    for index, (name, _) in enumerate(sorted_starts):
        next_row_by_name[name] = (
            sorted_starts[index + 1][1] - 1 if index + 1 < len(sorted_starts) else detail_ws.max_row
        )

    docs: list[ApiDoc] = []
    for name, title in interface_list:
        start_row = starts.get(name)
        if not start_row:
            continue

        end_row = next_row_by_name[name]
        method = "POST"
        path = ""
        summary = title
        params: list[dict[str, str]] = []
        array_details: dict[str, ArrayDetail] = {}
        current_section = ""
        current_array_name = ""

        for row in range(start_row, end_row + 1):
            first_col = safe_cell(detail_ws, row, 1)
            third_col = safe_cell(detail_ws, row, 3)
            fifth_col = safe_cell(detail_ws, row, 5)
            sixth_col = safe_cell(detail_ws, row, 6)

            if third_col == "公开URL":
                path = fifth_col
            elif third_col == "Http方法":
                method = fifth_col or method
            elif third_col == "命令概述":
                summary = fifth_col or summary

            if first_col == "参数":
                current_section = "params"
                continue
            if first_col == "数组类型参数详情":
                current_section = "array"
                current_array_name = third_col.replace("[数组类型参数名称]", "").strip()
                if current_array_name:
                    array_details[current_array_name] = ArrayDetail(name=current_array_name)
                continue
            if first_col in {"数据验证", "返回值", "命令"}:
                current_section = ""
                current_array_name = ""
                continue

            if current_section == "params" and third_col and third_col != "参数名称":
                params.append(
                    {
                        "name": third_col,
                        "type": fifth_col or "基础类型",
                        "required": "是",
                        "desc": sixth_col,
                    }
                )
                continue

            if current_section == "array" and current_array_name:
                if third_col and third_col not in {"[只发送变更数据]否", "[只发送变更数据]", "[主键列]", "数组项"}:
                    array_details[current_array_name].items.append(third_col)

        docs.append(
            ApiDoc(
                name=name,
                title=title,
                display_name=guess_display_name(name, title, path),
                method=method,
                path=path,
                summary=summary,
                params=params,
                array_details=array_details,
            )
        )

    return docs


def render_openapi(docs: list[ApiDoc]) -> str:
    paths: dict[str, dict] = {}

    for api in docs:
        request_properties = {}
        required = []
        example = {}

        for param in api.params:
            request_properties[param["name"]] = build_schema_for_param(api, param)
            required.append(param["name"])

            if param["type"] == "数组类型":
                detail = api.array_details.get(param["name"])
                if detail and detail.items:
                    example[param["name"]] = [{item_name: "" for item_name in detail.items}]
                else:
                    example[param["name"]] = []
            else:
                example[param["name"]] = ""

        normalized_path = api.path.replace("~", "").strip() or f"/ServerCommand/{api.name}"
        if not normalized_path.startswith("/"):
            normalized_path = f"/{normalized_path}"

        paths[normalized_path] = {
            api.method.lower(): {
                "tags": ["Excel Import"],
                "summary": api.display_name,
                "description": api.summary or api.title,
                "operationId": api.name,
                "requestBody": {
                    "required": True,
                    "content": {
                        "application/json": {
                            "schema": {
                                "type": "object",
                                "properties": request_properties,
                                "required": required,
                                "additionalProperties": False,
                            },
                            "example": example,
                        }
                    },
                },
                "responses": {
                    "200": {
                        "description": "成功",
                        "content": {
                            "application/json": {
                                "schema": {
                                    "type": "object",
                                    "description": "原始文档未提供明确返回结构",
                                    "additionalProperties": True,
                                }
                            }
                        },
                    }
                },
            }
        }

    spec = {
        "openapi": "3.0.3",
        "info": {
            "title": "Excel Imported API",
            "version": "1.0.0",
            "description": "从 Excel 接口说明文档提取生成。默认以详情 sheet 中每个接口首次出现的定义为准。",
        },
        "servers": [{"url": "/"}],
        "tags": [{"name": "Excel Import", "description": "从 Excel 提取的接口定义"}],
        "paths": paths,
    }
    return json.dumps(spec, ensure_ascii=False, indent=2) + "\n"


def default_output_path(source_xlsx: Path) -> Path:
    return source_xlsx.parent / "output" / "spreadsheet" / f"{source_xlsx.stem}-openapi.json"


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python3 excel_to_openapi.py <input.xlsx> [output.json]")
        sys.exit(1)

    source_xlsx = Path(sys.argv[1]).expanduser().resolve()
    output_json = Path(sys.argv[2]).expanduser().resolve() if len(sys.argv) > 2 else default_output_path(source_xlsx)

    docs = parse_interfaces(source_xlsx)
    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(render_openapi(docs), encoding="utf-8")

    print(f"generated: {output_json}")
    print(f"interfaces: {len(docs)}")


if __name__ == "__main__":
    main()
